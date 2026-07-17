from fastapi import APIRouter, Depends, HTTPException, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from database import get_db
from models import Queue, Participant, SwapRequest, QueueStatus, ParticipantStatus, LatecomerPolicy
from models.models import gen_4digit_code
import secrets
import asyncio
from datetime import datetime, timezone, timedelta
from io import BytesIO

router = APIRouter(prefix="/api")

_join_locks: dict[str, asyncio.Lock] = {}
_swap_locks: dict[str, asyncio.Lock] = {}


def calc_slot(start_time_str: str, slot_dur_min: int, group_idx: int):
    try:
        h, m = map(int, start_time_str.split(":"))
    except Exception:
        h, m = 9, 0
    base = datetime(2026, 1, 1, h, m)
    slot_start = base + timedelta(minutes=slot_dur_min * group_idx)
    slot_end = slot_start + timedelta(minutes=slot_dur_min)
    return slot_start.strftime("%H:%M"), slot_end.strftime("%H:%M")


async def get_slot_counts(db: AsyncSession, queue_id: str):
    result = await db.execute(
        select(Participant.slot_group, func.count()).where(
            Participant.queue_id == queue_id,
            Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
            Participant.slot_group.isnot(None),
        ).group_by(Participant.slot_group)
    )
    return dict(result.all())


async def find_next_slot(db: AsyncSession, queue_id: str, max_per_slot: int):
    counts = await get_slot_counts(db, queue_id)
    for i in range(200):
        if counts.get(i, 0) < max_per_slot:
            return i
    return 0


async def check_slot_capacity(db: AsyncSession, queue_id: str, slot_group, max_per_slot: int):
    counts = await get_slot_counts(db, queue_id)
    return counts.get(slot_group, 0) < max_per_slot


@router.post("/queue/create")
async def create_queue(
    name: str = Form(...),
    description: str = Form(""),
    max_size: int = Form(50),
    max_per_slot: int = Form(1),
    allow_swap: bool = Form(False),
    allow_delay: bool = Form(True),
    safe_window_minutes: int = Form(5),
    latecomer_policy: str = Form("to_end"),
    use_time_slots: bool = Form(True),
    slot_duration_minutes: int = Form(30),
    start_time: str = Form("09:00"),
    info: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    token = secrets.token_urlsafe(16)
    code = gen_4digit_code()
    while True:
        existing = await db.execute(select(Queue).where(Queue.code == code))
        if not existing.scalar_one_or_none():
            break
        code = gen_4digit_code()

    queue = Queue(
        code=code,
        name=name,
        description=description,
        organizer_token=token,
        max_size=max_size,
        max_per_slot=max(max_per_slot, 1),
        allow_swap=allow_swap,
        allow_delay=allow_delay,
        safe_window_minutes=safe_window_minutes,
        latecomer_policy=LatecomerPolicy(latecomer_policy),
        use_time_slots=use_time_slots,
        slot_duration_minutes=slot_duration_minutes,
        start_time=start_time,
        info=info if info.strip() else None,
    )
    db.add(queue)
    await db.commit()
    await db.refresh(queue)
    return {"queue_id": queue.id, "queue_code": queue.code, "organizer_token": token}


@router.get("/queue/{queue_id}")
async def get_queue(queue_id: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue:
        raise HTTPException(404, "Queue not found")
    use_slots = queue.use_time_slots if queue.use_time_slots is not None else True

    if use_slots:
        result = await db.execute(
            select(Participant).where(Participant.queue_id == queue_id).order_by(
                Participant.slot_group.asc(),
                Participant.position.asc(),
            )
        )
    else:
        result = await db.execute(
            select(Participant).where(Participant.queue_id == queue_id).order_by(Participant.position)
        )
    participants = result.scalars().all()

    slot_dur = queue.slot_duration_minutes or 30
    start_t = queue.start_time or "09:00"
    max_per = queue.max_per_slot or 1

    slot_data = []
    for p in participants:
        group = p.slot_group if p.slot_group is not None else 0
        if use_slots:
            slot_start, slot_end = calc_slot(start_t, slot_dur, group)
        else:
            slot_start, slot_end = None, None

        item = {
            "id": p.id,
            "name": p.name,
            "position": p.position,
            "slot_group": group,
            "status": p.status.value,
            "user_token": p.user_token,
            "joined_at": p.joined_at.isoformat(),
            "called_at": p.called_at.isoformat() if p.called_at else None,
            "served_at": p.served_at.isoformat() if p.served_at else None,
            "slot_start": slot_start,
            "slot_end": slot_end,
            "slot_label": None,
        }

        if not use_slots:
            if p.status == ParticipantStatus.WAITING or p.status == ParticipantStatus.DELAYED:
                item["slot_label"] = None
            elif p.status == ParticipantStatus.CALLED:
                item["slot_label"] = ""
            elif p.status == ParticipantStatus.SERVED:
                item["slot_label"] = ""
            elif p.status == ParticipantStatus.SKIPPED:
                item["slot_label"] = ""
        else:
            if p.status in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
                item["slot_label"] = f"{slot_start} – {slot_end}"
            elif p.status == ParticipantStatus.CALLED:
                item["slot_label"] = ""
            elif p.status == ParticipantStatus.SERVED:
                item["slot_label"] = ""
            elif p.status == ParticipantStatus.SKIPPED:
                item["slot_label"] = ""

        slot_data.append(item)

    swaps_result = await db.execute(
        select(SwapRequest).where(SwapRequest.queue_id == queue_id, SwapRequest.status == "pending")
    )
    pending_swaps = swaps_result.scalars().all()

    slot_counts = await get_slot_counts(db, queue_id)

    slots_info = []
    if use_slots:
        max_slot = max((slot_counts.keys() or {0})) if slot_counts else 0
        for i in range(max_slot + 2):
            ss, se = calc_slot(start_t, slot_dur, i)
            slots_info.append({
                "index": i,
                "start": ss,
                "end": se,
                "label": f"{ss} – {se}",
                "count": slot_counts.get(i, 0),
                "max": max_per,
            })

    return {
        "id": queue.id,
        "code": queue.code,
        "name": queue.name,
        "description": queue.description,
        "status": queue.status.value,
        "max_size": queue.max_size,
        "max_per_slot": max_per,
        "allow_swap": queue.allow_swap,
        "allow_delay": queue.allow_delay,
        "safe_window_minutes": queue.safe_window_minutes,
        "latecomer_policy": queue.latecomer_policy.value,
        "use_time_slots": use_slots,
        "slot_duration_minutes": slot_dur,
        "start_time": start_t,
        "info": queue.info,
        "created_at": queue.created_at.isoformat(),
        "slots": slots_info,
        "participants": slot_data,
        "participant_count": len(participants),
        "pending_swaps": [
            {
                "id": s.id,
                "from_id": s.from_participant_id,
                "from_name": s.from_participant.name if s.from_participant else "?",
                "to_id": s.to_participant_id,
                "to_name": s.to_participant.name if s.to_participant else None,
                "swap_type": s.swap_type,
                "target_slot_group": s.to_participant.slot_group if s.to_participant else s.target_slot_group,
                "target_slot_label": (
                    f"{calc_slot(start_t, slot_dur, s.to_participant.slot_group)[0]} – {calc_slot(start_t, slot_dur, s.to_participant.slot_group)[1]}"
                    if s.to_participant and s.to_participant.slot_group is not None else None
                ),
                "target_position": s.to_participant.position if s.to_participant else s.target_position,
            }
            for s in pending_swaps
        ],
    }


@router.get("/queue/by-code/{code}")
async def get_queue_by_code(code: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Queue).where(Queue.code == code))
    queue = result.scalar_one_or_none()
    if not queue:
        raise HTTPException(404, "Queue not found")
    return {"queue_id": queue.id, "code": queue.code, "name": queue.name}


@router.post("/queue/{queue_id}/join")
async def join_queue(
    queue_id: str,
    name: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    queue = await db.get(Queue, queue_id)
    if not queue:
        raise HTTPException(404, "Queue not found")
    if queue.status != QueueStatus.ACTIVE:
        raise HTTPException(400, "Queue is not active")

    if queue_id not in _join_locks:
        _join_locks[queue_id] = asyncio.Lock()

    async with _join_locks[queue_id]:
        count_result = await db.execute(
            select(func.count()).select_from(Participant).where(
                Participant.queue_id == queue_id,
                Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
            )
        )
        current_count = count_result.scalar()
        if current_count >= queue.max_size:
            raise HTTPException(400, "Queue is full")

        pos_result = await db.execute(
            select(func.max(Participant.position)).where(Participant.queue_id == queue_id)
        )
        max_pos = pos_result.scalar() or 0

        slot_group = None
        if queue.use_time_slots:
            slot_group = await find_next_slot(db, queue_id, queue.max_per_slot or 1)

        user_token = secrets.token_urlsafe(16)
        participant = Participant(
            queue_id=queue_id,
            name=name,
            position=max_pos + 1,
            slot_group=slot_group,
            status=ParticipantStatus.WAITING,
            user_token=user_token,
        )
        db.add(participant)
        await db.commit()
        await db.refresh(participant)
        return {"participant_id": participant.id, "user_token": user_token, "position": participant.position, "slot_group": slot_group}


@router.post("/queue/{queue_id}/admin/{token}/next")
async def next_in_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    if queue.status != QueueStatus.ACTIVE:
        raise HTTPException(400, "Queue is not active")

    result = await db.execute(
        select(Participant).where(
            Participant.queue_id == queue_id,
            Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.DELAYED]),
        ).order_by(Participant.position).limit(1)
    )
    current = result.scalar_one_or_none()
    if current:
        current.status = ParticipantStatus.CALLED
        current.called_at = datetime.now(timezone.utc)
        await db.commit()
        return {"called": {"id": current.id, "name": current.name, "position": current.position}}
    return {"called": None}


@router.post("/queue/{queue_id}/admin/{token}/serve/{participant_id}")
async def serve_participant(queue_id: str, token: str, participant_id: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    participant = await db.get(Participant, participant_id)
    if not participant or participant.queue_id != queue_id:
        raise HTTPException(404, "Participant not found")
    participant.status = ParticipantStatus.SERVED
    participant.served_at = datetime.now(timezone.utc)

    pos_result = await db.execute(
        select(Participant).where(
            Participant.queue_id == queue_id,
            Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
        ).order_by(Participant.position)
    )
    active = list(pos_result.scalars().all())
    for i, p in enumerate(active):
        p.position = i + 1

    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/admin/{token}/undo/{participant_id}")
async def undo_participant(queue_id: str, token: str, participant_id: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    participant = await db.get(Participant, participant_id)
    if not participant or participant.queue_id != queue_id:
        raise HTTPException(404, "Participant not found")
    if participant.status in (ParticipantStatus.SERVED, ParticipantStatus.SKIPPED):
        participant.status = ParticipantStatus.WAITING
        participant.served_at = None
        participant.called_at = None

        pos_result = await db.execute(
            select(Participant).where(
                Participant.queue_id == queue_id,
                Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
            ).order_by(Participant.position)
        )
        active = list(pos_result.scalars().all())
        for i, p in enumerate(active):
            p.position = i + 1

        await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/admin/{token}/skip/{participant_id}")
async def skip_participant(queue_id: str, token: str, participant_id: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    participant = await db.get(Participant, participant_id)
    if not participant or participant.queue_id != queue_id:
        raise HTTPException(404, "Participant not found")
    participant.status = ParticipantStatus.SKIPPED

    pos_result = await db.execute(
        select(Participant).where(
            Participant.queue_id == queue_id,
            Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
        ).order_by(Participant.position)
    )
    active = list(pos_result.scalars().all())
    for i, p in enumerate(active):
        p.position = i + 1

    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/admin/{token}/set-slot")
async def set_slot_group(
    queue_id: str, token: str,
    participant_id: str = Form(...),
    slot_group: int = Form(...),
    db: AsyncSession = Depends(get_db),
):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    participant = await db.get(Participant, participant_id)
    if not participant or participant.queue_id != queue_id:
        raise HTTPException(404, "Participant not found")
    has_space = await check_slot_capacity(db, queue_id, slot_group, queue.max_per_slot or 1)
    if not has_space:
        raise HTTPException(400, "Слот заполнен")
    participant.slot_group = slot_group
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/admin/{token}/set-slot-batch")
async def set_slot_batch(
    queue_id: str, token: str,
    participant_ids: str = Form(...),
    slot_group: int = Form(...),
    db: AsyncSession = Depends(get_db),
):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    ids = [x.strip() for x in participant_ids.split(",") if x.strip()]
    counts = await get_slot_counts(db, queue_id)
    current = counts.get(slot_group, 0)
    if current + len(ids) > (queue.max_per_slot or 1):
        raise HTTPException(400, f"Слот заполнен ({current}/{queue.max_per_slot})")
    count = 0
    for pid in ids:
        participant = await db.get(Participant, pid)
        if participant and participant.queue_id == queue_id:
            participant.slot_group = slot_group
            count += 1
    await db.commit()
    return {"ok": True, "updated": count}


@router.post("/queue/{queue_id}/admin/{token}/pause")
async def pause_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    queue.status = QueueStatus.PAUSED
    await db.commit()
    return {"status": "paused"}


@router.post("/queue/{queue_id}/admin/{token}/resume")
async def resume_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    queue.status = QueueStatus.ACTIVE
    await db.commit()
    return {"status": "active"}


@router.post("/queue/{queue_id}/admin/{token}/reopen")
async def reopen_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    if queue.status != QueueStatus.CLOSED:
        raise HTTPException(400, "Queue is not closed")
    queue.status = QueueStatus.ACTIVE
    await db.commit()
    return {"status": "active"}


@router.post("/queue/{queue_id}/admin/{token}/close")
async def close_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    queue.status = QueueStatus.CLOSED
    await db.commit()
    return {"status": "closed"}


@router.post("/queue/{queue_id}/admin/{token}/update-settings")
async def update_settings(
    queue_id: str, token: str,
    allow_swap: bool = Form(False),
    allow_delay: bool = Form(True),
    safe_window_minutes: int = Form(5),
    latecomer_policy: str = Form("to_end"),
    use_time_slots: bool = Form(True),
    slot_duration_minutes: int = Form(30),
    start_time: str = Form("09:00"),
    max_per_slot: int = Form(1),
    max_size: int = Form(50),
    info: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    queue.allow_swap = allow_swap
    queue.allow_delay = allow_delay
    queue.safe_window_minutes = safe_window_minutes
    queue.latecomer_policy = LatecomerPolicy(latecomer_policy)
    queue.slot_duration_minutes = slot_duration_minutes
    queue.start_time = start_time
    queue.max_per_slot = max(max_per_slot, 1)
    queue.max_size = max(max_size, 2)
    queue.info = info if info.strip() else None

    if queue.use_time_slots and not use_time_slots:
        result = await db.execute(
            select(Participant).where(
                Participant.queue_id == queue_id,
                Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
            )
        )
        for p in result.scalars().all():
            p.slot_group = None

    queue.use_time_slots = use_time_slots
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/leave")
async def leave_queue(queue_id: str, user_token: str = Form(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Participant).where(
            Participant.queue_id == queue_id,
            Participant.user_token == user_token,
            Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
        )
    )
    participant = result.scalar_one_or_none()
    if not participant:
        raise HTTPException(404, "Participant not found")
    participant.status = ParticipantStatus.SKIPPED
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/delay/{participant_id}")
async def mark_delayed(queue_id: str, participant_id: str, user_token: str = Form(...), db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue:
        raise HTTPException(404, "Queue not found")
    if queue.status != QueueStatus.ACTIVE:
        raise HTTPException(400, "Queue is not active")
    if not queue.allow_delay:
        raise HTTPException(400, "Delay is not allowed")
    participant = await db.get(Participant, participant_id)
    if not participant or participant.queue_id != queue_id or participant.user_token != user_token:
        raise HTTPException(403, "Unauthorized")
    participant.status = ParticipantStatus.DELAYED
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/swap/request")
async def request_swap(
    queue_id: str,
    from_id: str = Form(...),
    from_token: str = Form(...),
    to_id: str = Form(...),
    swap_type: str = Form("slot_change"),
    db: AsyncSession = Depends(get_db),
):
    queue = await db.get(Queue, queue_id)
    if not queue:
        raise HTTPException(404, "Queue not found")
    if queue.status != QueueStatus.ACTIVE:
        raise HTTPException(400, "Очередь не активна")
    if not queue.allow_swap:
        raise HTTPException(400, "Обмен не разрешён")
    from_p = await db.get(Participant, from_id)
    if not from_p or from_p.user_token != from_token:
        raise HTTPException(403, "Unauthorized")
    if from_p.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
        raise HTTPException(400, "Нельзя менять статус")

    to_p = await db.get(Participant, to_id)
    if not to_p or to_p.queue_id != queue_id:
        raise HTTPException(404, "Участник не найден")
    if to_p.id == from_p.id:
        raise HTTPException(400, "Нельзя меняться с самим собой")
    if to_p.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
        raise HTTPException(400, "Участник недоступен для обмена")

    existing = await db.execute(
        select(SwapRequest).where(
            SwapRequest.queue_id == queue_id,
            SwapRequest.from_participant_id == from_id,
            SwapRequest.status == "pending",
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, "У вас уже есть активный запрос")

    if swap_type == "slot_change":
        if not queue.use_time_slots:
            raise HTTPException(400, "Тайм-слоты отключены")
        if from_p.slot_group == to_p.slot_group:
            raise HTTPException(400, "Вы уже в одном слоте")
    elif swap_type == "position_change":
        if from_p.slot_group != to_p.slot_group:
            raise HTTPException(400, "Вы в разных слотах")

    lock = _swap_locks.setdefault(queue_id, asyncio.Lock())
    async with lock:
        mutual = await db.execute(
            select(SwapRequest).where(
                SwapRequest.queue_id == queue_id,
                SwapRequest.from_participant_id == to_id,
                SwapRequest.to_participant_id == from_id,
                SwapRequest.swap_type == swap_type,
                SwapRequest.status == "pending",
            )
        )
        mutual_swap = mutual.scalar_one_or_none()

        if mutual_swap:
            from_p_db = await db.get(Participant, from_id)
            to_p_db = await db.get(Participant, to_id)
            if not from_p_db or not to_p_db:
                raise HTTPException(404, "Участник не найден")
            if from_p_db.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED) or \
               to_p_db.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
                mutual_swap.status = "rejected"
                new_swap = SwapRequest(
                    queue_id=queue_id,
                    from_participant_id=from_id,
                    to_participant_id=to_id,
                    swap_type=swap_type,
                    status="rejected",
                )
                db.add(new_swap)
                await db.commit()
                return {"swap_id": new_swap.id, "status": "rejected"}

            if swap_type == "slot_change":
                from_p_db.slot_group, to_p_db.slot_group = to_p_db.slot_group, from_p_db.slot_group
            elif swap_type == "position_change":
                from_p_db.position, to_p_db.position = to_p_db.position, from_p_db.position

            mutual_swap.status = "auto_accepted"

            new_swap = SwapRequest(
                queue_id=queue_id,
                from_participant_id=from_id,
                to_participant_id=to_id,
                swap_type=swap_type,
                status="auto_accepted",
            )
            db.add(new_swap)

            pos_result = await db.execute(
                select(Participant).where(
                    Participant.queue_id == queue_id,
                    Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
                ).order_by(Participant.position)
            )
            active = list(pos_result.scalars().all())
            for i, p in enumerate(active):
                p.position = i + 1

            await db.commit()
            return {"swap_id": new_swap.id, "status": "auto_accepted"}

        swap = SwapRequest(
            queue_id=queue_id,
            from_participant_id=from_id,
            to_participant_id=to_id,
            swap_type=swap_type,
            status="pending",
        )
        db.add(swap)
        await db.commit()
        await db.refresh(swap)
        return {"swap_id": swap.id, "status": "pending"}


@router.post("/queue/{queue_id}/swap/cancel/{swap_id}")
async def cancel_swap(queue_id: str, swap_id: str, user_token: str = Form(...), db: AsyncSession = Depends(get_db)):
    swap = await db.get(SwapRequest, swap_id)
    if not swap or swap.queue_id != queue_id:
        raise HTTPException(404, "Запрос не найден")
    from_p = await db.get(Participant, swap.from_participant_id)
    if not from_p or from_p.user_token != user_token:
        raise HTTPException(403, "Unauthorized")
    if swap.status != "pending":
        raise HTTPException(400, "Запрос уже обработан")
    swap.status = "cancelled"
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/swap/accept/{swap_id}")
async def participant_accept_swap(queue_id: str, swap_id: str, user_token: str = Form(...), db: AsyncSession = Depends(get_db)):
    swap = await db.get(SwapRequest, swap_id)
    if not swap or swap.queue_id != queue_id or swap.status != "pending":
        raise HTTPException(404, "Запрос не найден")
    to_p = await db.get(Participant, swap.to_participant_id)
    if not to_p or to_p.user_token != user_token:
        raise HTTPException(403, "Unauthorized")
    if to_p.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
        raise HTTPException(400, "Невозможно выполнить обмен")

    from_p = await db.get(Participant, swap.from_participant_id)
    if not from_p:
        raise HTTPException(404, "Участник не найден")
    if from_p.status not in (ParticipantStatus.WAITING, ParticipantStatus.DELAYED):
        raise HTTPException(400, "Участник недоступен")

    queue = await db.get(Queue, queue_id)

    lock = _swap_locks.setdefault(queue_id, asyncio.Lock())
    async with lock:
        if swap.swap_type == "slot_change":
            if not queue.use_time_slots:
                raise HTTPException(400, "Тайм-слоты отключены")
            from_p.slot_group, to_p.slot_group = to_p.slot_group, from_p.slot_group
        elif swap.swap_type == "position_change":
            from_p.position, to_p.position = to_p.position, from_p.position

        swap.status = "accepted"

        other_incoming = await db.execute(
            select(SwapRequest).where(
                SwapRequest.queue_id == queue_id,
                SwapRequest.to_participant_id == to_p.id,
                SwapRequest.status == "pending",
                SwapRequest.id != swap.id,
            )
        )
        for other in other_incoming.scalars().all():
            other.status = "rejected"

        pos_result = await db.execute(
            select(Participant).where(
                Participant.queue_id == queue_id,
                Participant.status.in_([ParticipantStatus.WAITING, ParticipantStatus.CALLED, ParticipantStatus.DELAYED]),
            ).order_by(Participant.position)
        )
        active = list(pos_result.scalars().all())
        for i, p in enumerate(active):
            p.position = i + 1

        await db.commit()
        return {"ok": True}


@router.post("/queue/{queue_id}/swap/reject/{swap_id}")
async def participant_reject_swap(queue_id: str, swap_id: str, user_token: str = Form(...), db: AsyncSession = Depends(get_db)):
    swap = await db.get(SwapRequest, swap_id)
    if not swap or swap.queue_id != queue_id:
        raise HTTPException(404, "Запрос не найден")
    to_p = await db.get(Participant, swap.to_participant_id)
    if not to_p or to_p.user_token != user_token:
        raise HTTPException(403, "Unauthorized")
    if swap.status != "pending":
        raise HTTPException(400, "Запрос уже обработан")
    swap.status = "rejected"
    await db.commit()
    return {"ok": True}


@router.post("/queue/{queue_id}/admin/{token}/delete")
async def delete_queue(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")
    await db.delete(queue)
    await db.commit()
    return {"ok": True}


@router.get("/queue/{queue_id}/admin/{token}/export")
async def export_participants(queue_id: str, token: str, db: AsyncSession = Depends(get_db)):
    queue = await db.get(Queue, queue_id)
    if not queue or queue.organizer_token != token:
        raise HTTPException(403, "Unauthorized")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    result = await db.execute(
        select(Participant).where(Participant.queue_id == queue_id).order_by(Participant.position)
    )
    participants = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = queue.name[:31] if queue.name else "Очередь"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_map = {
        "waiting": "Ожидает",
        "called": "Вызван",
        "served": "Обслужён",
        "delayed": "Задерживается",
        "skipped": "Пропущен",
    }

    slot_dur = queue.slot_duration_minutes or 30
    start_t = queue.start_time or "09:00"
    use_slots = queue.use_time_slots if queue.use_time_slots is not None else True

    headers = ["№", "Имя", "Статус", "Позиция"]
    if use_slots:
        headers.extend(["Слот", "Время слота"])
    headers.extend(["Записался", "Вызван", "Обслужён"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for idx, p in enumerate(participants, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=p.name).border = thin_border
        status_cell = ws.cell(row=row, column=3, value=status_map.get(p.status.value, p.status.value))
        status_cell.border = thin_border
        ws.cell(row=row, column=4, value=p.position).border = thin_border

        col_offset = 5
        if use_slots:
            slot_group = p.slot_group if p.slot_group is not None else 0
            ss, se = calc_slot(start_t, slot_dur, slot_group)
            ws.cell(row=row, column=col_offset, value=slot_group + 1).border = thin_border
            ws.cell(row=row, column=col_offset + 1, value=f"{ss} – {se}").border = thin_border
            col_offset += 2

        joined_str = p.joined_at.strftime("%d.%m.%Y %H:%M") if p.joined_at else ""
        called_str = p.called_at.strftime("%d.%m.%Y %H:%M") if p.called_at else ""
        served_str = p.served_at.strftime("%d.%m.%Y %H:%M") if p.served_at else ""
        ws.cell(row=row, column=col_offset, value=joined_str).border = thin_border
        ws.cell(row=row, column=col_offset + 1, value=called_str).border = thin_border
        ws.cell(row=row, column=col_offset + 2, value=served_str).border = thin_border

        status_colors = {
            "waiting": "FFF2CC",
            "called": "D9E2F3",
            "served": "E2EFDA",
            "delayed": "FCE4D6",
            "skipped": "D9D9D9",
        }
        color = status_colors.get(p.status.value)
        if color:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"queue_{queue.code}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
